import openpyxl


class Excel:

    def __init__(self, log):
        self.log = log

    def get_input(self, inputFile):
        
        # GET INPUT FROM EXCEL SECTION #
        self.log.logging_info(f'Getting Inputs from {inputFile}')
        
        try:
            wb = openpyxl.load_workbook(inputFile)
        except FileNotFoundError:
            self.log.logging_error(f"The file {inputFile} was not found.")
        except Exception as e:
            self.log.logging_error(f"Error loading workbook: {e}")

        wb_obj = openpyxl.load_workbook(inputFile)

        sheet_obj = wb_obj.active

        self.desc = sheet_obj.cell(row = 2, column = 1).value
        self.name1 = sheet_obj.cell(row = 2, column = 2).value
        self.descrip = sheet_obj.cell(row = 2, column = 3).value
        self.emailID = sheet_obj.cell(row = 2, column = 4).value
        self.firstName = sheet_obj.cell(row = 2, column = 5).value
        self.groupName = sheet_obj.cell(row = 2, column = 6).value
        self.lastName = sheet_obj.cell(row = 2, column = 7).value
        self.password = sheet_obj.cell(row = 2, column = 8).value
        self.userName = sheet_obj.cell(row = 2, column = 9).value
        self.description = sheet_obj.cell(row = 2, column = 10).value
        self.enabled = sheet_obj.cell(row = 2, column = 11).value
        self.name2 = sheet_obj.cell(row = 2, column = 12).value

        bSuccess=True
        if (self.desc is None):
            self.log.logging_error("Error: Desc is missing")
            bSuccess=False
        if (self.name1 is None):
            self.log.logging_error("Error: Name1 is missing")
            bSuccess=False
        if (self.descrip is None):
            self.log.logging_error("Error: Descrip is missing")
            bSuccess=False
        if (self.emailID is None):
            self.log.logging_error("Error: EmailID is missing")
            bSuccess=False
        if (self.firstName is None):
            self.log.logging_error("Error: FirstName is missing")
            bSuccess=False
        if (self.groupName is None):
            self.log.logging_error("Error: Group Name is missing")
            bSuccess=False
        if (self.lastName is None):
            self.log.logging_error("Error: Last Name is missing")
            bSuccess=False
        if (self.password is None):
            self.log.logging_error("Error: Password is missing")
            bSuccess=False
        if (self.userName is None):
            self.log.logging_error("Error: User Name is missing")
            bSuccess=False
        if (self.description is None):
            self.log.logging_error("Error: Description is missing")
            bSuccess=False
        if (self.enabled is None):
            self.log.logging_error("Error: Number is missing")
            bSuccess=False
        if (self.name2 is None):
            self.log.logging_error("Error: Name2 is missing")
            bSuccess=False

        return bSuccess
